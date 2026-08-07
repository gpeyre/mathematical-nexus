#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from common import HERE, ROOT, URL_RE, build_alt_text, build_comment, load_config, schedule_path


COLUMNS = [
    "queue_id", "batch_id", "batch_position", "filename", "media_type", "title",
    "comment", "alt_text", "publication_date", "publication_time", "timezone",
    "scheduled_at_local", "scheduled_at_utc", "status", "record_key",
    "posted_at_utc", "post_uri", "post_cid", "last_error",
]


def previous_values(path: Path) -> dict[str, dict]:
    if not path.exists():
        return {}
    wb = load_workbook(path, data_only=False)
    ws = wb["schedule"]
    headers = [c.value for c in ws[1]]
    return {
        str(row[headers.index("filename")]): dict(zip(headers, row))
        for row in ws.iter_rows(min_row=2, values_only=True)
    }


def source_urls() -> dict[str, list[str]]:
    source = (ROOT / "vignettes" / "mydata.js").read_text(encoding="utf-8")
    marker = "const textData = `"
    if marker not in source:
        return {}
    body = source.split(marker, 1)[1].rsplit("`", 1)[0]
    result: dict[str, list[str]] = {}
    for block in body.split("\n\n"):
        lines = [line.strip() for line in block.splitlines() if line.strip()]
        if len(lines) < 2:
            continue
        urls = [url.rstrip(".,;)") for url in URL_RE.findall(" ".join(lines[2:]))]
        if urls:
            result[lines[1]] = urls
    return result


def style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="16324F")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    widths = {
        "A": 11, "B": 10, "C": 14, "D": 38, "E": 12, "F": 30, "G": 76,
        "H": 76, "I": 16, "J": 16, "K": 18, "L": 28, "M": 24, "N": 15,
        "O": 28, "P": 24, "Q": 55, "R": 55, "S": 55,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        row[6].alignment = Alignment(wrap_text=True, vertical="top")
        row[7].alignment = Alignment(wrap_text=True, vertical="top")


def write_workbook(path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "schedule"
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row.get(column, "") for column in COLUMNS])
    style_sheet(ws)

    info = wb.create_sheet("instructions")
    info.append(["Mathematical Nexus Bluesky schedule"])
    info.append(["Edit comment, alt_text, or publication date/time before approving a batch."])
    info.append(["Statuses: review -> approved -> posted. The API runner publishes only approved due rows."])
    info.append(["Bluesky has no native future-post scheduling endpoint; this workbook is the local queue."])
    info.column_dimensions["A"].width = 110
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the Bluesky vignette schedule.")
    parser.add_argument("--start-date", help="YYYY-MM-DD; defaults to config.json")
    parser.add_argument("--force-comments", action="store_true", help="Regenerate edited comments and alt text")
    args = parser.parse_args()

    config = load_config()
    tz = ZoneInfo(config["timezone"])
    start = date.fromisoformat(args.start_date or config["start_date"])
    hour, minute = map(int, config["publication_time"].split(":"))
    batch_size = int(config["batch_size"])
    max_graphemes = int(config["max_graphemes"])
    db_path = (HERE / config["source_database"]).resolve()
    catalog = json.loads(db_path.read_text(encoding="utf-8"))
    entries = sorted(
        (row for row in catalog if row.get("type") == "vignette" and (ROOT / row["filename"]).exists()),
        key=lambda row: row["filename"],
    )
    old = previous_values(schedule_path(config))
    original_urls = source_urls()

    rows: list[dict] = []
    for index, entry in enumerate(entries, start=1):
        local_dt = datetime.combine(start + timedelta(days=index - 1), time(hour, minute), tzinfo=tz)
        utc_dt = local_dt.astimezone(timezone.utc)
        filename = entry["filename"]
        suffix = Path(filename).suffix.lower()
        media_type = "video" if suffix in {".m4v", ".mp4"} else "image"
        prior = old.get(filename, {})
        preferred_url = (original_urls.get(Path(filename).name) or [""])[0]
        comment = build_comment(entry["title"], entry["content"], max_graphemes, preferred_url)
        alt_text = build_alt_text(entry["title"], entry["content"])
        if prior and not args.force_comments:
            comment = prior.get("comment") or comment
            alt_text = prior.get("alt_text") or alt_text
        batch_id = math.ceil(index / batch_size)
        row = {
            "queue_id": index,
            "batch_id": batch_id,
            "batch_position": (index - 1) % batch_size + 1,
            "filename": filename,
            "media_type": media_type,
            "title": entry["title"],
            "comment": comment,
            "alt_text": alt_text,
            "publication_date": local_dt.date().isoformat(),
            "publication_time": local_dt.strftime("%H:%M"),
            "timezone": config["timezone"],
            "scheduled_at_local": local_dt.isoformat(timespec="minutes"),
            "scheduled_at_utc": utc_dt.isoformat(timespec="minutes").replace("+00:00", "Z"),
            "status": prior.get("status", "review") if prior else "review",
            "record_key": f"mn-{index:04d}-{local_dt:%Y%m%d}",
            "posted_at_utc": prior.get("posted_at_utc", "") if prior else "",
            "post_uri": prior.get("post_uri", "") if prior else "",
            "post_cid": prior.get("post_cid", "") if prior else "",
            "last_error": prior.get("last_error", "") if prior else "",
        }
        rows.append(row)

    output = schedule_path(config)
    write_workbook(output, rows)
    batch_dir = HERE / "batches"
    batch_dir.mkdir(exist_ok=True)
    for batch_id in sorted({row["batch_id"] for row in rows}):
        batch_rows = [row for row in rows if row["batch_id"] == batch_id]
        write_workbook(batch_dir / f"batch_{batch_id:03d}.xlsx", batch_rows)
        (batch_dir / f"batch_{batch_id:03d}.json").write_text(
            json.dumps(batch_rows, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    print(f"Wrote {len(rows)} scheduled vignettes to {output}")
    print(f"Wrote {math.ceil(len(rows) / batch_size)} review batches to {batch_dir}")


if __name__ == "__main__":
    main()
