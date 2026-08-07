#!/usr/bin/env python3
from __future__ import annotations

import argparse
import subprocess
import sys
from datetime import date, datetime, time, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from common import HERE, read_schedule_rows, schedule_path, set_row_values


EDITABLE_COLUMNS = {"title", "comment", "alt_text", "publication_date", "publication_time", "timezone"}


def iso_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def iso_time(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value)[:5]


def main() -> None:
    parser = argparse.ArgumentParser(description="Import editorial changes from one batch workbook.")
    parser.add_argument("--batch", required=True, type=int)
    args = parser.parse_args()
    batch_path = HERE / "batches" / f"batch_{args.batch:03d}.xlsx"
    if not batch_path.exists():
        raise SystemExit(f"Missing {batch_path}")

    batch_wb = load_workbook(batch_path, data_only=False)
    batch_ws = batch_wb["schedule"]
    batch_headers = [cell.value for cell in batch_ws[1]]
    batch_rows = [dict(zip(batch_headers, row)) for row in batch_ws.iter_rows(min_row=2, values_only=True)]

    master_wb, master_ws, master_rows = read_schedule_rows()
    master_by_filename = {str(row["filename"]): row for row in master_rows}
    changed = 0
    for batch_row in batch_rows:
        filename = str(batch_row["filename"])
        master = master_by_filename.get(filename)
        if not master or int(master["batch_id"]) != args.batch:
            raise SystemExit(f"Batch row does not match master schedule: {filename}")
        publication_date = iso_date(batch_row["publication_date"])
        publication_time = iso_time(batch_row["publication_time"])
        timezone_name = str(batch_row["timezone"])
        hour, minute = map(int, publication_time.split(":"))
        local_dt = datetime.combine(date.fromisoformat(publication_date), time(hour, minute), tzinfo=ZoneInfo(timezone_name))
        utc_dt = local_dt.astimezone(timezone.utc)
        updates = {column: batch_row[column] for column in EDITABLE_COLUMNS if column in batch_row}
        updates.update({
            "publication_date": publication_date,
            "publication_time": publication_time,
            "scheduled_at_local": local_dt.isoformat(timespec="minutes"),
            "scheduled_at_utc": utc_dt.isoformat(timespec="minutes").replace("+00:00", "Z"),
            "record_key": f"mn-{int(master['queue_id']):04d}-{local_dt:%Y%m%d}",
        })
        set_row_values(master_ws, master["_excel_row"], updates)
        changed += 1
    master_wb.save(schedule_path())
    subprocess.run([sys.executable, "validate_schedule.py", "--batch", str(args.batch)], check=True)
    print(f"Imported {changed} reviewed rows from {batch_path}")


if __name__ == "__main__":
    main()
