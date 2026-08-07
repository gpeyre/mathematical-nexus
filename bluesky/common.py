from __future__ import annotations

import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

import regex
from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
URL_RE = re.compile(r"https?://[^\s]+")


def load_config() -> dict[str, Any]:
    return json.loads((HERE / "config.json").read_text(encoding="utf-8"))


def schedule_path(config: dict[str, Any] | None = None) -> Path:
    override = os.environ.get("BLUESKY_SCHEDULE_PATH")
    if override:
        return Path(override).expanduser().resolve()
    config = config or load_config()
    return HERE / config["schedule_file"]


def load_dotenv(path: Path | None = None) -> None:
    path = path or HERE / ".env"
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def grapheme_len(text: str) -> int:
    return len(regex.findall(r"\X", text))


def truncate_graphemes(text: str, limit: int) -> str:
    parts = regex.findall(r"\X", text)
    if len(parts) <= limit:
        return text
    return "".join(parts[: max(0, limit - 1)]).rstrip() + "…"


def clean_sentence(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def usable_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.hostname) and "." in parsed.hostname


def build_comment(title: str, content: str, limit: int = 300, preferred_url: str = "") -> str:
    urls = [u.rstrip(".,;)") for u in URL_RE.findall(content)]
    prose = clean_sentence(URL_RE.sub("", content))
    base = clean_sentence(f"{title} — {prose}" if prose else title)
    candidates = [preferred_url.rstrip(".,;)")] if preferred_url else []
    candidates.extend(urls)
    url = next((value for value in candidates if usable_url(value)), "")
    if url:
        room = limit - grapheme_len(url) - 1
        if room >= 30:
            base = truncate_graphemes(base, room)
            return f"{base} {url}"
    return truncate_graphemes(base, limit)


def build_alt_text(title: str, content: str, limit: int = 1000) -> str:
    prose = clean_sentence(URL_RE.sub("", content))
    return truncate_graphemes(clean_sentence(f"{title}. {prose}"), limit)


def parse_local_timestamp(value: str, timezone_name: str) -> datetime:
    dt = datetime.fromisoformat(value)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=ZoneInfo(timezone_name))
    return dt


def read_schedule_rows(path: Path | None = None) -> tuple[Any, Any, list[dict[str, Any]]]:
    path = path or schedule_path()
    workbook = load_workbook(path)
    sheet = workbook["schedule"]
    headers = [cell.value for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for excel_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(headers, values))
        row["_excel_row"] = excel_row
        rows.append(row)
    return workbook, sheet, rows


def set_row_values(sheet: Any, excel_row: int, updates: dict[str, Any]) -> None:
    headers = {cell.value: cell.column for cell in sheet[1]}
    for key, value in updates.items():
        if key in headers:
            sheet.cell(excel_row, headers[key], value)
